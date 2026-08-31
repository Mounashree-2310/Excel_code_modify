import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelReport:
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file

    def generate_report(self):
        temp_df = pd.read_excel(
            self.input_file,
            sheet_name="Report",
            header=None,
        )

        header_row = None

        for i in range(len(temp_df)):
            if "CATEGORY" in temp_df.iloc[i].astype(str).tolist():
                header_row = i
                break

        df = pd.read_excel(
            self.input_file,
            sheet_name="Report",
            header=header_row,
        )

        df.columns = df.columns.astype(str).str.strip()

        filtered_df = df[
            df["CATEGORY"].astype(str).str.strip()
            == "IRREGULAR_TEST_FAILURE"
        ]

        filtered_df["TICKET"] = (
            filtered_df["TICKET"]
            .fillna("NULL")
            .replace("", "NULL")
        )

        filtered_df.to_excel(
            self.output_file,
            index=False,
        )

        wb = load_workbook(self.output_file)
        ws = wb.active

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)

            cell.fill = PatternFill(
                start_color="D9EAF7",
                end_color="D9EAF7",
                fill_type="solid",
            )

        # Wrap text
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            ws.column_dimensions[column_letter].width = min(
                max_length + 3,
                50,
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(self.output_file)

        return len(filtered_df)


def main():
    report = ExcelReport(
        "failuresClassifications_MGU22_03-08-2026-10-08-2026.xlsx",
        "Irregular_Test_Failures.xlsx",
    )

    count = report.generate_report()

    print("Report Generated Successfully")
    print(f"Total Irregular Failures: {count}")


if __name__ == "__main__":
    main()
